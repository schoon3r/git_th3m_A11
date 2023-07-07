#!/usr/bin/env python3

import os
import sys
from bs4 import BeautifulSoup
from openpyxl import Workbook

# Check if the html_directory argument is provided
if len(sys.argv) < 2:
    print("Please provide the directory containing the HTML files.")
    sys.exit(1)

# Get the html_directory from the command-line argument
html_directory = sys.argv[1]

# Create a new workbook
workbook = Workbook()

# Iterate over the HTML files
for filename in os.listdir(html_directory):
    if filename.endswith(".html"):
        file_path = os.path.join(html_directory, filename)

        # Read the HTML file
        with open(file_path, "r") as file:
            html_content = file.read()

        # Parse the HTML using BeautifulSoup
        soup = BeautifulSoup(html_content, "html.parser")

        # Find the table in the HTML content
        table = soup.find("table")

        # Create a new worksheet for the HTML file
        worksheet = workbook.create_sheet(title=filename)

        # Write the table content to the worksheet
        for row_index, row in enumerate(table.find_all("tr")):
            for col_index, cell in enumerate(row.find_all(["th", "td"])):
                worksheet.cell(row=row_index + 1, column=col_index +
                               1, value=cell.get_text(strip=True))

# Remove the default sheet created by openpyxl
workbook.remove(workbook["Sheet"])

# Save the workbook as combined.xlsx
workbook.save("combined.xlsx")
